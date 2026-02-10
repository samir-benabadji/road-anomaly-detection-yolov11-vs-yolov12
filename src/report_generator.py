import shutil
from pathlib import Path
from datetime import datetime

def generate_report(report_info: dict, test_metrics: dict,
                    results_dir: Path, train_run_dir: Path) -> list:
    model_name = report_info["model_name"]
    mode = report_info["mode"]
    report_base = f"report_{model_name}_{mode}"

    generated = []

    _copy_training_graphs(train_run_dir, results_dir)

    try:
        csv_path = train_run_dir / "results.csv"
        if csv_path.exists():
            epoch_graph_path = results_dir / f"epoch_curves_{model_name}_{mode}.png"
            _generate_epoch_curves(csv_path, model_name, mode, epoch_graph_path)
            generated.append(str(epoch_graph_path))
    except Exception as e:
        print(f"  Could not generate epoch curves: {e}")

    txt_path = results_dir / f"{report_base}.txt"
    _generate_txt_report(report_info, test_metrics, txt_path)
    generated.append(str(txt_path))

    try:
        xlsx_path = results_dir / f"{report_base}.xlsx"
        _generate_excel_report(report_info, test_metrics, xlsx_path)
        generated.append(str(xlsx_path))
    except ImportError:
        print("  openpyxl not installed, skipping Excel report.")
        print("  Install with: pip install openpyxl")

    try:
        chart_path = results_dir / f"metrics_summary_{model_name}_{mode}.png"
        _generate_metrics_chart(test_metrics, model_name, mode, chart_path)
        generated.append(str(chart_path))
    except Exception as e:
        print(f"  Could not generate chart: {e}")

    return generated

def _copy_training_graphs(train_run_dir: Path, results_dir: Path):
    graphs_dir = results_dir / "graphs"
    graphs_dir.mkdir(exist_ok=True)

    graph_files = [
        "results.png", "confusion_matrix.png",
        "confusion_matrix_normalized.png",
        "F1_curve.png", "P_curve.png", "R_curve.png",
        "PR_curve.png", "labels.jpg", "labels_correlogram.jpg",
    ]

    for fname in graph_files:
        src = train_run_dir / fname
        if src.exists():
            shutil.copy2(src, graphs_dir / fname)

def _generate_txt_report(report_info: dict, metrics: dict, path: Path):
    model = report_info["model_name"]
    mode = report_info["mode"]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines = []
    lines.append("=" * 65)
    lines.append(f"  TRAINING REPORT - {model} ({mode})")
    lines.append(f"  Road Anomaly Detection - Projet Fin d'Etude - Master 2")
    lines.append("=" * 65)
    lines.append("")

    lines.append("-" * 65)
    lines.append("  TRAINING CONFIGURATION")
    lines.append("-" * 65)
    lines.append(f"  Date               : {now}")
    lines.append(f"  Model              : {model}")
    lines.append(f"  Variant            : {report_info.get('model_variant', 'N/A')}")
    lines.append(f"  Training Mode      : {mode}")
    lines.append(f"  Epochs             : {report_info.get('epochs', 'N/A')}")
    lines.append(f"  Batch Size         : {report_info.get('batch_size', 'N/A')}")
    lines.append(f"  Image Size         : {report_info.get('image_size', 640)}")
    lines.append(f"  Learning Rate (lr0): {report_info.get('lr0', 'N/A')}")
    lines.append(f"  Final LR (lrf)     : {report_info.get('lrf', 'N/A')}")
    lines.append(f"  Optimizer          : {report_info.get('optimizer', 'N/A')}")
    lines.append(f"  Cosine LR Schedule : {report_info.get('cos_lr', False)}")
    if "freeze_layers" in report_info:
        lines.append(f"  Frozen Layers      : {report_info['freeze_layers']}")
    lines.append(f"  Device             : {report_info.get('device', 'N/A')}")
    lines.append(f"  GPU                : {report_info.get('gpu_name', 'N/A')}")
    lines.append(f"  Training Time      : {report_info.get('training_time', 'N/A')}")
    lines.append("")

    lines.append("-" * 65)
    lines.append("  TEST SET METRICS")
    lines.append("-" * 65)

    if metrics:
        lines.append(f"  {'Metric':<30s}  {'Value':>10s}")
        lines.append(f"  {'------':<30s}  {'-----':>10s}")

        key_metrics = [
            ("mAP@0.5", "mAP50"),
            ("mAP@0.5:0.95", "mAP50-95"),
            ("Precision", "precision"),
            ("Recall", "recall"),
            ("F1 Score", "f1_score"),
            ("Model Size (MB)", "model_size_mb"),
        ]

        for display, key in key_metrics:
            val = metrics.get(key)
            if val is not None:
                if "size" in key.lower():
                    lines.append(f"  {display:<30s}  {val:>10.2f}")
                else:
                    lines.append(f"  {display:<30s}  {val:>10.4f}")
    else:
        lines.append("  (No test metrics available)")

    lines.append("")

    per_class = {k: v for k, v in metrics.items() if k.startswith("AP50_")}
    if per_class:
        lines.append("-" * 65)
        lines.append("  PER-CLASS AVERAGE PRECISION (AP@0.5)")
        lines.append("-" * 65)
        lines.append(f"  {'Class':<35s}  {'AP@0.5':>10s}")
        lines.append(f"  {'-----':<35s}  {'------':>10s}")
        for k, v in per_class.items():
            cls_name = k.replace("AP50_", "")
            lines.append(f"  {cls_name:<35s}  {v:>10.4f}")
        lines.append("")

    lines.append("-" * 65)
    lines.append("  TRAINING MODE EXPLANATION")
    lines.append("-" * 65)
    if mode == "Standard":
        lines.append("  Standard training uses default YOLO hyperparameters:")
        lines.append("  - All model layers are trainable (no frozen layers)")
        lines.append("  - Linear learning rate decay schedule")
        lines.append("  - SGD optimizer (auto-selected by YOLO)")
        lines.append(f"  - Initial learning rate: {report_info.get('lr0', 0.01)}")
    else:
        lines.append("  Improved training uses Freeze Backbone + Cosine LR:")
        lines.append(f"  - First {report_info.get('freeze_layers', 10)} backbone layers are FROZEN")
        lines.append("    (preserves pre-trained feature extraction)")
        lines.append("  - Cosine annealing LR schedule (smooth decay)")
        lines.append("  - AdamW optimizer (better for fine-tuning)")
        lines.append(f"  - Smaller initial learning rate: {report_info.get('lr0', 0.001)}")
        lines.append("  - Longer warmup period for stable convergence")
    lines.append("")

    lines.append("=" * 65)
    lines.append(f"  Report generated: {now}")
    lines.append("=" * 65)

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"  Saved: {path.name}")

def _generate_excel_report(report_info: dict, metrics: dict, path: Path):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Training Config"

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    ws1.merge_cells('A1:B1')
    ws1['A1'] = f"Training Report - {report_info['model_name']} ({report_info['mode']})"
    ws1['A1'].font = Font(bold=True, size=14)

    ws1.merge_cells('A2:B2')
    ws1['A2'] = "Road Anomaly Detection - Projet Fin d'Etude - Master 2"
    ws1['A2'].font = Font(italic=True, size=10)

    ws1.merge_cells('A3:B3')
    ws1['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    for col, header in [(1, "Parameter"), (2, "Value")]:
        cell = ws1.cell(row=5, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    config_rows = [
        ("Model", report_info["model_name"]),
        ("Variant", report_info.get("model_variant", "N/A")),
        ("Training Mode", report_info["mode"]),
        ("Epochs", report_info.get("epochs", "N/A")),
        ("Batch Size", report_info.get("batch_size", "N/A")),
        ("Image Size", report_info.get("image_size", 640)),
        ("Learning Rate (lr0)", report_info.get("lr0", "N/A")),
        ("Final LR ratio (lrf)", report_info.get("lrf", "N/A")),
        ("Optimizer", report_info.get("optimizer", "N/A")),
        ("Cosine LR Schedule", str(report_info.get("cos_lr", False))),
        ("Device", report_info.get("device", "N/A")),
        ("GPU", report_info.get("gpu_name", "N/A")),
        ("Training Time", report_info.get("training_time", "N/A")),
    ]

    if "freeze_layers" in report_info:
        config_rows.insert(3, ("Frozen Layers", report_info["freeze_layers"]))

    for i, (param, value) in enumerate(config_rows, start=6):
        ws1.cell(row=i, column=1, value=param).border = thin_border
        ws1.cell(row=i, column=2, value=str(value)).border = thin_border

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 40

    ws2 = wb.create_sheet("Test Metrics")

    ws2.merge_cells('A1:B1')
    ws2['A1'] = "Test Set Evaluation Metrics"
    ws2['A1'].font = Font(bold=True, size=14)

    for col, header in [(1, "Metric"), (2, "Value")]:
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    if metrics:
        metric_rows = [
            ("mAP@0.5", metrics.get("mAP50")),
            ("mAP@0.5:0.95", metrics.get("mAP50-95")),
            ("Precision", metrics.get("precision")),
            ("Recall", metrics.get("recall")),
            ("F1 Score", metrics.get("f1_score")),
            ("Model Size (MB)", metrics.get("model_size_mb")),
        ]

        for i, (name, val) in enumerate(metric_rows, start=4):
            ws2.cell(row=i, column=1, value=name).border = thin_border
            cell = ws2.cell(row=i, column=2)
            cell.border = thin_border
            if val is not None:
                cell.value = round(val, 4) if val < 10 else round(val, 2)
                cell.number_format = '0.0000' if val < 10 else '0.00'

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        for row in range(4, 4 + len(metric_rows)):
            val = ws2.cell(row=row, column=2).value
            if val is not None and isinstance(val, (int, float)) and val > 0.5:
                ws2.cell(row=row, column=2).fill = green_fill

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 20

    per_class = {k: v for k, v in metrics.items() if k.startswith("AP50_")}
    if per_class:
        ws3 = wb.create_sheet("Per-Class AP")

        ws3.merge_cells('A1:B1')
        ws3['A1'] = "Per-Class Average Precision (AP@0.5)"
        ws3['A1'].font = Font(bold=True, size=14)

        for col, header in [(1, "Class"), (2, "AP@0.5")]:
            cell = ws3.cell(row=3, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for i, (k, v) in enumerate(per_class.items(), start=4):
            cls_name = k.replace("AP50_", "")
            ws3.cell(row=i, column=1, value=cls_name).border = thin_border
            cell = ws3.cell(row=i, column=2, value=round(v, 4))
            cell.border = thin_border
            cell.number_format = '0.0000'

        ws3.column_dimensions['A'].width = 35
        ws3.column_dimensions['B'].width = 15

    ws4 = wb.create_sheet("Method Info")

    ws4.merge_cells('A1:A1')
    ws4['A1'] = "Training Method Explanation"
    ws4['A1'].font = Font(bold=True, size=14)

    mode = report_info["mode"]
    if mode == "Standard":
        explanations = [
            "Standard training uses default YOLO hyperparameters:",
            "- All model layers are trainable (no frozen layers)",
            "- Linear learning rate decay schedule",
            "- SGD optimizer (auto-selected by YOLO)",
            f"- Initial learning rate: {report_info.get('lr0', 0.01)}",
            "",
            "This serves as the baseline for comparison.",
        ]
    else:
        explanations = [
            "Improved training: Freeze Backbone + Cosine LR Schedule",
            "",
            "Technique details:",
            f"1. Frozen Backbone: First {report_info.get('freeze_layers', 10)} layers are frozen",
            "   - These layers contain pre-trained feature extractors",
            "   - Freezing preserves learned features from ImageNet",
            "   - Only the detection head is trained on our data",
            "",
            "2. Cosine Annealing LR: Smooth learning rate decay",
            "   - LR follows a cosine curve from lr0 to lr0*lrf",
            "   - Avoids sharp drops, helps convergence",
            "",
            "3. AdamW Optimizer: Better for fine-tuning tasks",
            "   - Adaptive learning rates per parameter",
            "   - Weight decay applied correctly (decoupled)",
            "",
            f"4. Smaller Initial LR: {report_info.get('lr0', 0.001)}",
            "   - Prevents overshooting optimal weights",
            "   - More stable when using frozen backbone",
        ]

    for i, line in enumerate(explanations, start=3):
        ws4.cell(row=i, column=1, value=line)

    ws4.column_dimensions['A'].width = 60

    wb.save(path)
    print(f"  Saved: {path.name}")

def _generate_epoch_curves(csv_path: Path, model_name: str, mode: str, path: Path):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import csv

    epochs, map50s, map50_95s = [], [], []
    precisions, recalls = [], []
    box_losses, cls_losses, dfl_losses = [], [], []
    val_box_losses, val_cls_losses = [], []

    with open(csv_path, 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            cleaned = {k.strip(): v.strip() for k, v in row.items()}
            ep = int(cleaned.get('epoch', 0))
            epochs.append(ep)
            map50s.append(float(cleaned.get('metrics/mAP50(B)', 0)))
            map50_95s.append(float(cleaned.get('metrics/mAP50-95(B)', 0)))
            precisions.append(float(cleaned.get('metrics/precision(B)', 0)))
            recalls.append(float(cleaned.get('metrics/recall(B)', 0)))
            box_losses.append(float(cleaned.get('train/box_loss', 0)))
            cls_losses.append(float(cleaned.get('train/cls_loss', 0)))
            dfl_losses.append(float(cleaned.get('train/dfl_loss', 0)))
            val_box_losses.append(float(cleaned.get('val/box_loss', 0)))
            val_cls_losses.append(float(cleaned.get('val/cls_loss', 0)))

    if not epochs:
        return

    best_idx = map50s.index(max(map50s))
    best_epoch = epochs[best_idx]
    best_map50 = map50s[best_idx]

    fig, axes = plt.subplots(2, 2, figsize=(18, 12))
    fig.suptitle(f'{model_name} ({mode}) - Epoch Training Curves\n'
                 f'Best Epoch: {best_epoch} (mAP50 = {best_map50:.4f})',
                 fontsize=16, fontweight='bold', y=0.98)

    ax1 = axes[0, 0]
    ax1.plot(epochs, map50s, 'b-o', markersize=3, linewidth=1.5, label='mAP@0.5', alpha=0.8)
    ax1.plot(epochs, map50_95s, 'r-o', markersize=3, linewidth=1.5, label='mAP@0.5:0.95', alpha=0.8)
    ax1.plot(best_epoch, best_map50, '*', color='gold', markersize=20,
             markeredgecolor='black', markeredgewidth=1, label=f'Best (epoch {best_epoch})', zorder=5)
    ax1.axvline(x=best_epoch, color='gold', linestyle='--', alpha=0.5)
    ax1.set_xlabel('Epoch', fontsize=11)
    ax1.set_ylabel('Score', fontsize=11)
    ax1.set_title('mAP per Epoch', fontsize=13, fontweight='bold')
    ax1.legend(fontsize=10)
    ax1.grid(True, alpha=0.3)
    ax1.set_ylim(0, max(max(map50s) * 1.15, 0.1))

    ax2 = axes[0, 1]
    ax2.plot(epochs, precisions, 'g-o', markersize=3, linewidth=1.5, label='Precision', alpha=0.8)
    ax2.plot(epochs, recalls, 'm-o', markersize=3, linewidth=1.5, label='Recall', alpha=0.8)
    f1s = [2 * (p * r) / (p + r + 1e-6) for p, r in zip(precisions, recalls)]
    ax2.plot(epochs, f1s, 'c-o', markersize=3, linewidth=1.5, label='F1 Score', alpha=0.8)
    ax2.axvline(x=best_epoch, color='gold', linestyle='--', alpha=0.5, label=f'Best epoch ({best_epoch})')
    ax2.set_xlabel('Epoch', fontsize=11)
    ax2.set_ylabel('Score', fontsize=11)
    ax2.set_title('Precision / Recall / F1 per Epoch', fontsize=13, fontweight='bold')
    ax2.legend(fontsize=10)
    ax2.grid(True, alpha=0.3)
    ax2.set_ylim(0, 1.05)

    ax3 = axes[1, 0]
    ax3.plot(epochs, box_losses, 'r-o', markersize=2, linewidth=1.2, label='Box Loss', alpha=0.8)
    ax3.plot(epochs, cls_losses, 'b-o', markersize=2, linewidth=1.2, label='Cls Loss', alpha=0.8)
    ax3.plot(epochs, dfl_losses, 'g-o', markersize=2, linewidth=1.2, label='DFL Loss', alpha=0.8)
    ax3.axvline(x=best_epoch, color='gold', linestyle='--', alpha=0.5, label=f'Best epoch ({best_epoch})')
    ax3.set_xlabel('Epoch', fontsize=11)
    ax3.set_ylabel('Loss', fontsize=11)
    ax3.set_title('Training Losses per Epoch', fontsize=13, fontweight='bold')
    ax3.legend(fontsize=10)
    ax3.grid(True, alpha=0.3)

    ax4 = axes[1, 1]
    ax4.plot(epochs, val_box_losses, 'r-o', markersize=2, linewidth=1.2, label='Val Box Loss', alpha=0.8)
    ax4.plot(epochs, val_cls_losses, 'b-o', markersize=2, linewidth=1.2, label='Val Cls Loss', alpha=0.8)
    ax4.axvline(x=best_epoch, color='gold', linestyle='--', alpha=0.5, label=f'Best epoch ({best_epoch})')
    ax4.set_xlabel('Epoch', fontsize=11)
    ax4.set_ylabel('Loss', fontsize=11)
    ax4.set_title('Validation Losses per Epoch', fontsize=13, fontweight='bold')
    ax4.legend(fontsize=10)
    ax4.grid(True, alpha=0.3)

    plt.tight_layout(rect=[0, 0, 1, 0.95])
    fig.savefig(path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  Saved: {path.name}")

def _generate_metrics_chart(metrics: dict, model_name: str, mode: str, path: Path):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np

    if not metrics:
        return

    metric_names = ["mAP@0.5", "mAP@0.5:0.95", "Precision", "Recall", "F1 Score"]
    metric_keys = ["mAP50", "mAP50-95", "precision", "recall", "f1_score"]
    values = [metrics.get(k, 0) for k in metric_keys]

    fig, ax = plt.subplots(figsize=(10, 6))

    colors = ['#2196F3', '#4CAF50', '#FF9800', '#9C27B0', '#F44336']
    bars = ax.bar(metric_names, values, color=colors, edgecolor='black', linewidth=0.8)

    for bar, val in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.01,
                f'{val:.4f}', ha='center', va='bottom', fontsize=11, fontweight='bold')

    ax.set_ylim(0, 1.05)
    ax.set_ylabel('Score', fontsize=12)
    ax.set_title(f'{model_name} ({mode}) - Test Set Metrics', fontsize=14, fontweight='bold')
    ax.grid(axis='y', alpha=0.3)
    ax.set_axisbelow(True)

    plt.tight_layout()
    fig.savefig(path, dpi=200, bbox_inches='tight')
    plt.close()
    print(f"  Saved: {path.name}")
