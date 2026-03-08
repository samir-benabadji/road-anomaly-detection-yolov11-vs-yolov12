import os
import sys
import shutil
import csv
from pathlib import Path
from datetime import datetime
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).parent.parent))
from config import DATA_DIR, DATASET_CONFIG
from src.utils import print_colored

def setup_percountry_test_dirs():
    countries = DATASET_CONFIG["countries"]
    test_images_dir = DATA_DIR / "test" / "images"
    test_labels_dir = DATA_DIR / "test" / "labels"

    if not test_images_dir.exists():
        print_colored("  Test directory not found! Run setup_dataset.py first.", "red")
        return False

    created_countries = []

    for country in countries:
        country_test_dir = DATA_DIR / f"test_{country}"
        country_img_dir = country_test_dir / "images"
        country_lbl_dir = country_test_dir / "labels"

        if country_img_dir.exists() and len(list(country_img_dir.iterdir())) > 0:
            created_countries.append(country)
            continue

        country_img_dir.mkdir(parents=True, exist_ok=True)
        country_lbl_dir.mkdir(parents=True, exist_ok=True)

        count = 0
        for img_file in test_images_dir.iterdir():
            if not img_file.suffix.lower() in ('.jpg', '.jpeg', '.png'):
                continue
            if not img_file.name.startswith(country + "_"):
                continue

            dst_img = country_img_dir / img_file.name
            if not dst_img.exists():
                try:
                    os.link(str(img_file), str(dst_img))
                except (OSError, NotImplementedError):
                    shutil.copy2(img_file, dst_img)

            lbl_file = test_labels_dir / (img_file.stem + ".txt")
            dst_lbl = country_lbl_dir / lbl_file.name
            if lbl_file.exists() and not dst_lbl.exists():
                try:
                    os.link(str(lbl_file), str(dst_lbl))
                except (OSError, NotImplementedError):
                    shutil.copy2(lbl_file, dst_lbl)

            count += 1

        yaml_path = DATA_DIR / f"data_{country}_test.yaml"
        yaml_content = f"""# Per-Country Test Configuration - {country}
# Train/Val: All countries combined | Test: {country} only

path: {DATA_DIR}
train: train/images
val: val/images
test: test_{country}/images

nc: {DATASET_CONFIG['num_classes']}

names:
  0: D00_Longitudinal_Crack
  1: D10_Transverse_Crack
  2: D20_Alligator_Crack
  3: D40_Pothole
"""
        with open(yaml_path, 'w') as f:
            f.write(yaml_content)

        print(f"    {country}: {count} test images")
        created_countries.append(country)

    return len(created_countries) == len(countries)

def evaluate_percountry(model_path: str, results_dir: Path, device: str = "0",
                        use_tta: bool = False):
    from ultralytics import YOLO

    countries = DATASET_CONFIG["countries"]
    class_names = DATASET_CONFIG["classes"]
    all_country_metrics = {}

    for country in countries:
        yaml_path = DATA_DIR / f"data_{country}_test.yaml"
        if not yaml_path.exists():
            print_colored(f"  Skipping {country}: yaml not found", "yellow")
            continue

        print_colored(f"\n  Evaluating on {country} test set...", "blue")

        country_eval_dir = results_dir / f"test_{country}"
        model = YOLO(str(model_path))

        val_kwargs = {
            "data": str(yaml_path),
            "split": "test",
            "device": device,
            "project": str(results_dir),
            "name": f"test_{country}",
            "exist_ok": True,
            "verbose": True,
        }
        if use_tta:
            val_kwargs["augment"] = True
        test_results = model.val(**val_kwargs)

        metrics = {
            "mAP50": float(test_results.box.map50),
            "mAP50-95": float(test_results.box.map),
            "precision": float(test_results.box.mp),
            "recall": float(test_results.box.mr),
        }
        p, r = metrics["precision"], metrics["recall"]
        metrics["f1_score"] = 2 * (p * r) / (p + r + 1e-6)

        if hasattr(test_results.box, 'ap50'):
            for i, cname in enumerate(class_names):
                if i < len(test_results.box.ap50):
                    metrics[f"AP50_{cname}"] = float(test_results.box.ap50[i])

        all_country_metrics[country] = metrics

        print(f"    {country}: mAP50={metrics['mAP50']:.4f}, "
              f"P={metrics['precision']:.4f}, R={metrics['recall']:.4f}, "
              f"F1={metrics['f1_score']:.4f}")

    return all_country_metrics

def generate_percountry_report(country_metrics: dict, overall_metrics: dict,
                                report_info: dict, results_dir: Path):
    model_name = report_info["model_name"]
    mode = report_info["mode"]
    variant = report_info.get("model_variant", "")
    generated = []

    txt_path = results_dir / f"percountry_report_{model_name}_{mode}.txt"
    _generate_percountry_txt(country_metrics, overall_metrics, report_info, txt_path)
    generated.append(str(txt_path))

    try:
        xlsx_path = results_dir / f"percountry_report_{model_name}_{mode}.xlsx"
        _generate_percountry_excel(country_metrics, overall_metrics, report_info, xlsx_path)
        generated.append(str(xlsx_path))
    except ImportError:
        print("  openpyxl not installed, skipping Excel report.")

    try:
        chart1_path = results_dir / f"percountry_mAP50_{model_name}_{mode}.png"
        _generate_country_bar_chart(country_metrics, overall_metrics,
                                    model_name, mode, chart1_path)
        generated.append(str(chart1_path))

        chart2_path = results_dir / f"percountry_class_heatmap_{model_name}_{mode}.png"
        _generate_class_country_heatmap(country_metrics, model_name, mode, chart2_path)
        generated.append(str(chart2_path))

        chart3_path = results_dir / f"percountry_all_metrics_{model_name}_{mode}.png"
        _generate_country_metrics_comparison(country_metrics, model_name, mode, chart3_path)
        generated.append(str(chart3_path))

        chart4_path = results_dir / f"percountry_radar_{model_name}_{mode}.png"
        _generate_country_radar(country_metrics, model_name, mode, chart4_path)
        generated.append(str(chart4_path))

    except Exception as e:
        print(f"  Could not generate per-country charts: {e}")
        import traceback
        traceback.print_exc()

    return generated

def _generate_percountry_txt(country_metrics, overall_metrics, report_info, path):
    model = report_info["model_name"]
    mode = report_info["mode"]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines = []
    lines.append("=" * 75)
    lines.append(f"  PER-COUNTRY TEST REPORT - {model} ({mode})")
    lines.append(f"  Road Anomaly Detection - Projet Fin d'Etude - Master 2")
    lines.append("=" * 75)
    lines.append("")

    lines.append("-" * 75)
    lines.append("  TRAINING CONFIGURATION")
    lines.append("-" * 75)
    lines.append(f"  Model: {model} ({report_info.get('model_variant', '')})")
    lines.append(f"  Mode: {mode}")
    lines.append(f"  Training: ALL countries combined")
    lines.append(f"  Validation: ALL countries combined")
    lines.append(f"  Testing: SEPARATE per country")
    lines.append(f"  Training Time: {report_info.get('training_time', 'N/A')}")
    lines.append(f"  GPU: {report_info.get('gpu_name', 'N/A')}")
    lines.append("")

    lines.append("-" * 75)
    lines.append("  OVERALL TEST METRICS (All Countries Combined)")
    lines.append("-" * 75)
    if overall_metrics:
        lines.append(f"  mAP@0.5     : {overall_metrics.get('mAP50', 0):.4f}")
        lines.append(f"  mAP@0.5:0.95: {overall_metrics.get('mAP50-95', 0):.4f}")
        lines.append(f"  Precision   : {overall_metrics.get('precision', 0):.4f}")
        lines.append(f"  Recall      : {overall_metrics.get('recall', 0):.4f}")
        lines.append(f"  F1 Score    : {overall_metrics.get('f1_score', 0):.4f}")
    lines.append("")

    lines.append("-" * 75)
    lines.append("  PER-COUNTRY TEST RESULTS")
    lines.append("-" * 75)
    header = f"  {'Country':<18s} {'mAP50':>8s} {'mAP50-95':>10s} {'Precision':>10s} {'Recall':>8s} {'F1':>8s}"
    lines.append(header)
    lines.append(f"  {'-------':<18s} {'-----':>8s} {'--------':>10s} {'---------':>10s} {'------':>8s} {'--':>8s}")

    for country, m in country_metrics.items():
        lines.append(f"  {country:<18s} {m['mAP50']:>8.4f} {m['mAP50-95']:>10.4f} "
                      f"{m['precision']:>10.4f} {m['recall']:>8.4f} {m['f1_score']:>8.4f}")
    lines.append("")

    lines.append("-" * 75)
    lines.append("  PER-CLASS AP@0.5 BY COUNTRY")
    lines.append("-" * 75)
    class_names = DATASET_CONFIG["classes"]
    header = f"  {'Country':<18s}" + "".join(f" {c[:12]:>13s}" for c in class_names)
    lines.append(header)
    lines.append(f"  {'-------':<18s}" + "".join(f" {'--------':>13s}" for _ in class_names))

    for country, m in country_metrics.items():
        vals = []
        for cname in class_names:
            v = m.get(f"AP50_{cname}", 0)
            vals.append(f" {v:>13.4f}")
        lines.append(f"  {country:<18s}" + "".join(vals))
    lines.append("")

    lines.append("-" * 75)
    lines.append("  ANALYSIS NOTES")
    lines.append("-" * 75)

    if country_metrics:
        best_country = max(country_metrics, key=lambda c: country_metrics[c]['mAP50'])
        worst_country = min(country_metrics, key=lambda c: country_metrics[c]['mAP50'])
        lines.append(f"  Best performing country  : {best_country} "
                      f"(mAP50 = {country_metrics[best_country]['mAP50']:.4f})")
        lines.append(f"  Worst performing country : {worst_country} "
                      f"(mAP50 = {country_metrics[worst_country]['mAP50']:.4f})")
        gap = country_metrics[best_country]['mAP50'] - country_metrics[worst_country]['mAP50']
        lines.append(f"  Performance gap          : {gap:.4f} ({gap*100:.1f}%)")
        lines.append("")
        lines.append("  Discussion: Differences in detection rates across countries may be due to:")
        lines.append("  - Image capture conditions (lighting, weather, camera angle)")
        lines.append("  - Road surface types and textures")
        lines.append("  - Image resolution differences across datasets")
        lines.append("  - Regional variation in road damage appearance")

    lines.append("")
    lines.append("=" * 75)
    lines.append(f"  Report generated: {now}")
    lines.append("=" * 75)

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  Saved: {path.name}")

def _generate_percountry_excel(country_metrics, overall_metrics, report_info, path):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    model = report_info["model_name"]
    mode = report_info["mode"]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws = wb.active
    ws.title = "Per-Country Results"

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Per-Country Test Results - {model} ({mode})"
    ws['A1'].font = Font(bold=True, size=14)

    headers = ["Country", "mAP@0.5", "mAP@0.5:0.95", "Precision", "Recall", "F1 Score"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    best_map50 = max(m['mAP50'] for m in country_metrics.values()) if country_metrics else 0
    worst_map50 = min(m['mAP50'] for m in country_metrics.values()) if country_metrics else 0

    for i, (country, m) in enumerate(country_metrics.items(), start=4):
        ws.cell(row=i, column=1, value=country).border = thin_border
        for j, key in enumerate(['mAP50', 'mAP50-95', 'precision', 'recall', 'f1_score'], 2):
            cell = ws.cell(row=i, column=j, value=round(m.get(key, 0), 4))
            cell.border = thin_border
            cell.number_format = '0.0000'
            if key == 'mAP50':
                if m['mAP50'] == best_map50:
                    cell.fill = green_fill
                elif m['mAP50'] == worst_map50:
                    cell.fill = red_fill

    row = len(country_metrics) + 4
    ws.cell(row=row, column=1, value="ALL (Combined)").border = thin_border
    ws.cell(row=row, column=1).font = Font(bold=True)
    if overall_metrics:
        for j, key in enumerate(['mAP50', 'mAP50-95', 'precision', 'recall', 'f1_score'], 2):
            cell = ws.cell(row=row, column=j, value=round(overall_metrics.get(key, 0), 4))
            cell.border = thin_border
            cell.number_format = '0.0000'
            cell.font = Font(bold=True)

    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 18

    ws2 = wb.create_sheet("Per-Class by Country")
    class_names = DATASET_CONFIG["classes"]

    ws2.merge_cells(f'A1:{chr(65 + len(class_names))}1')
    ws2['A1'] = f"Per-Class AP@0.5 by Country - {model} ({mode})"
    ws2['A1'].font = Font(bold=True, size=14)

    headers2 = ["Country"] + [c.replace("_", " ") for c in class_names]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, (country, m) in enumerate(country_metrics.items(), start=4):
        ws2.cell(row=i, column=1, value=country).border = thin_border
        for j, cname in enumerate(class_names, 2):
            val = m.get(f"AP50_{cname}", 0)
            cell = ws2.cell(row=i, column=j, value=round(val, 4))
            cell.border = thin_border
            cell.number_format = '0.0000'

    ws2.column_dimensions['A'].width = 18
    for col in range(2, 2 + len(class_names)):
        ws2.column_dimensions[chr(64 + col)].width = 22

    wb.save(path)
    print(f"  Saved: {path.name}")

def _generate_country_bar_chart(country_metrics, overall_metrics, model_name, mode, path):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np

    countries = list(country_metrics.keys())
    map50s = [country_metrics[c]['mAP50'] for c in countries]

    countries_all = countries + ["ALL\n(Combined)"]
    map50s_all = map50s + [overall_metrics.get('mAP50', 0)]

    fig, ax = plt.subplots(figsize=(12, 7))

    colors = ['#2196F3', '#4CAF50', '#FF9800', '#9C27B0', '#F44336', '#607D8B']
    bars = ax.bar(countries_all, map50s_all, color=colors[:len(countries_all)],
                  edgecolor='black', linewidth=0.8)

    for bar, val in zip(bars, map50s_all):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.005,
                f'{val:.4f}', ha='center', va='bottom', fontsize=11, fontweight='bold')

    best_idx = map50s.index(max(map50s))
    worst_idx = map50s.index(min(map50s))
    bars[best_idx].set_edgecolor('green')
    bars[best_idx].set_linewidth(3)
    bars[worst_idx].set_edgecolor('red')
    bars[worst_idx].set_linewidth(3)

    ax.set_ylim(0, max(map50s_all) * 1.15)
    ax.set_ylabel('mAP@0.5', fontsize=13)
    ax.set_title(f'{model_name} ({mode}) - mAP@0.5 per Country\n'
                 f'Train: All Combined | Test: Per Country',
                 fontsize=14, fontweight='bold')
    ax.grid(axis='y', alpha=0.3)
    ax.set_axisbelow(True)

    plt.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  Saved: {path.name}")

def _generate_class_country_heatmap(country_metrics, model_name, mode, path):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np

    class_names = DATASET_CONFIG["classes"]
    countries = list(country_metrics.keys())

    data = []
    for country in countries:
        row = []
        for cname in class_names:
            row.append(country_metrics[country].get(f"AP50_{cname}", 0))
        data.append(row)
    data = np.array(data)

    fig, ax = plt.subplots(figsize=(12, 7))

    im = ax.imshow(data, cmap='RdYlGn', aspect='auto', vmin=0, vmax=1)

    short_classes = ["D00\nLongitudinal", "D10\nTransverse", "D20\nAlligator", "D40\nPothole"]
    ax.set_xticks(range(len(class_names)))
    ax.set_xticklabels(short_classes, fontsize=11)
    ax.set_yticks(range(len(countries)))
    ax.set_yticklabels(countries, fontsize=12)

    for i in range(len(countries)):
        for j in range(len(class_names)):
            val = data[i, j]
            color = 'white' if val < 0.3 or val > 0.7 else 'black'
            ax.text(j, i, f'{val:.3f}', ha='center', va='center',
                    fontsize=13, fontweight='bold', color=color)

    ax.set_title(f'{model_name} ({mode}) - Per-Class AP@0.5 by Country',
                 fontsize=14, fontweight='bold', pad=15)

    cbar = plt.colorbar(im, ax=ax, shrink=0.8)
    cbar.set_label('AP@0.5', fontsize=12)

    plt.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  Saved: {path.name}")

def _generate_country_metrics_comparison(country_metrics, model_name, mode, path):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np

    countries = list(country_metrics.keys())
    metrics_to_show = ['mAP50', 'mAP50-95', 'precision', 'recall', 'f1_score']
    metric_labels = ['mAP@0.5', 'mAP@0.5:0.95', 'Precision', 'Recall', 'F1 Score']

    x = np.arange(len(countries))
    width = 0.15
    colors = ['#2196F3', '#4CAF50', '#FF9800', '#9C27B0', '#F44336']

    fig, ax = plt.subplots(figsize=(14, 7))

    for i, (metric_key, metric_label) in enumerate(zip(metrics_to_show, metric_labels)):
        values = [country_metrics[c].get(metric_key, 0) for c in countries]
        offset = (i - len(metrics_to_show) / 2 + 0.5) * width
        bars = ax.bar(x + offset, values, width, label=metric_label,
                      color=colors[i], edgecolor='black', linewidth=0.5)

    ax.set_ylabel('Score', fontsize=12)
    ax.set_title(f'{model_name} ({mode}) - All Metrics per Country',
                 fontsize=14, fontweight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(countries, fontsize=11)
    ax.legend(fontsize=10, loc='upper right')
    ax.set_ylim(0, 1.05)
    ax.grid(axis='y', alpha=0.3)
    ax.set_axisbelow(True)

    plt.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  Saved: {path.name}")

def _generate_country_radar(country_metrics, model_name, mode, path):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np

    countries = list(country_metrics.keys())
    metrics_keys = ['mAP50', 'precision', 'recall', 'f1_score']
    metric_labels = ['mAP@0.5', 'Precision', 'Recall', 'F1 Score']
    num_metrics = len(metrics_keys)

    angles = np.linspace(0, 2 * np.pi, num_metrics, endpoint=False).tolist()
    angles += angles[:1]

    fig, ax = plt.subplots(figsize=(10, 10), subplot_kw=dict(polar=True))

    colors = ['#2196F3', '#4CAF50', '#FF9800', '#9C27B0', '#F44336']

    for i, country in enumerate(countries):
        values = [country_metrics[country].get(k, 0) for k in metrics_keys]
        values += values[:1]
        ax.plot(angles, values, 'o-', linewidth=2, label=country,
                color=colors[i % len(colors)], markersize=6)
        ax.fill(angles, values, alpha=0.1, color=colors[i % len(colors)])

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(metric_labels, fontsize=12)
    ax.set_ylim(0, 1)
    ax.set_title(f'{model_name} ({mode}) - Country Comparison Radar',
                 fontsize=14, fontweight='bold', pad=20)
    ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1), fontsize=11)
    ax.grid(True)

    plt.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  Saved: {path.name}")
